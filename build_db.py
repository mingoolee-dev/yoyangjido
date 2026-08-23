"""
공단 공개데이터 → SQLite

원본: 국민건강보험공단_장기요양기관 시설별 현황 (기준일 2026-06-10)
      공공데이터포털 https://www.data.go.kr/data/15124763/fileData.do
      이용허락범위: 제한 없음 / 무료

사용법: python3 build_db.py <원본.xlsx> [시군구이름]
"""
import sys, re, sqlite3, unicodedata
from collections import defaultdict
import openpyxl

DATA_기준일 = "2026-06-10"
DATA_출처 = "국민건강보험공단 장기요양기관 시설별 현황 (공공데이터포털)"
DATA_URL = "https://www.data.go.kr/data/15124763/fileData.do"

# 원본 유형명 → 보호자가 아는 말 (12-5장: 사용자가 아는 말로)
유형맵 = [
    ("노인요양시설",        "요양원"),
    ("노인전문요양시설",     "요양원"),
    ("노인요양공동생활가정", "공동생활가정"),
    ("주야간보호",          "주야간보호"),
    ("단기보호",            "단기보호"),
    ("방문요양",            "방문요양"),
    ("방문목욕",            "방문목욕"),
    ("방문간호",            "방문간호"),
    ("복지용구",            "복지용구"),
    ("치매전담",            "치매전담실"),
]
유형순서 = ["요양원", "공동생활가정", "치매전담실", "주야간보호", "단기보호",
            "방문요양", "방문목욕", "방문간호", "복지용구"]


def 유형정리(원본명: str) -> str | None:
    if not 원본명:
        return None
    for key, label in 유형맵:
        if key in 원본명:
            return label
    return None


def slugify(name: str, code: str) -> str:
    """한글 시설명을 URL로. 한글을 그대로 두되 공백/기호만 정리하고, 뒤에 코드 4자리를 붙여 충돌 방지."""
    s = unicodedata.normalize("NFC", name or "").strip()
    s = re.sub(r"[\s/\\?#%&+]+", "-", s)
    s = re.sub(r"[^0-9A-Za-z가-힣()\-]", "", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return f"{s}-{code[-4:]}" if s else code


def main(xlsx_path: str, 시군구: str = "군산"):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    기관 = {}
    for i, r in enumerate(wb["일반현황"].iter_rows(values_only=True)):
        if i == 0 or not r[0]:
            continue
        지역 = r[6] or ""
        if 시군구 not in 지역:
            continue
        code = str(r[0])
        parts = 지역.split()
        기관[code] = {
            "code": code,
            "name": (r[1] or "").strip(),
            "sido": parts[0] if parts else "",
            "sigungu": parts[1] if len(parts) > 1 else "",
            "dong": parts[2] if len(parts) > 2 else "",
            "지정일자": str(r[7] or ""),
            "주소": (r[9] or "").strip(),
        }

    정원 = defaultdict(dict)
    유형 = defaultdict(set)
    for i, r in enumerate(wb["입소인원"].iter_rows(values_only=True)):
        if i == 0 or not r[0]:
            continue
        code = str(r[0])
        if code not in 기관:
            continue
        label = 유형정리(r[2])
        if not label:
            continue
        유형[code].add(label)
        if r[3]:
            정원[code][label] = max(정원[code].get(label, 0), int(r[3]))

    인력 = {}
    인력열 = ["시설장", "사무국장", "사회복지사", "의사_전임", "의사_촉탁", "간호사",
              "간호조무사", "치위생사", "물리치료사", "작업치료사", "요양보호사", "영양사", "기타"]
    for i, r in enumerate(wb["인력현황"].iter_rows(values_only=True)):
        if i == 0 or not r[0]:
            continue
        code = str(r[0])
        if code not in 기관:
            continue
        label = 유형정리(r[2])
        if not label:
            continue
        cur = 인력.setdefault(code, {k: 0 for k in 인력열})
        for j, k in enumerate(인력열):
            v = r[3 + j]
            if isinstance(v, (int, float)):
                cur[k] = max(cur[k], int(v))
    wb.close()

    con = sqlite3.connect("yoyangjido.db")
    con.executescript("""
        DROP TABLE IF EXISTS facility;
        DROP TABLE IF EXISTS facility_type;
        CREATE TABLE facility(
            code TEXT PRIMARY KEY, slug TEXT UNIQUE, name TEXT,
            sido TEXT, sigungu TEXT, dong TEXT,
            addr TEXT, desig_date TEXT,
            요양보호사 INTEGER, 간호사 INTEGER, 간호조무사 INTEGER,
            사회복지사 INTEGER, 물리치료사 INTEGER, 작업치료사 INTEGER,
            의사_촉탁 INTEGER, 영양사 INTEGER,
            인력합계 INTEGER, 대표유형 TEXT, 대표정원 INTEGER
        );
        CREATE TABLE facility_type(
            code TEXT, type TEXT, capacity INTEGER,
            PRIMARY KEY(code, type)
        );
        CREATE INDEX idx_sgg ON facility(sigungu);
        CREATE INDEX idx_type ON facility_type(type);
    """)

    used = set()
    rows, trows = [], []
    for code, f in 기관.items():
        ts = 유형[code]
        if not ts:
            continue  # 유형을 모르는 기관은 페이지를 만들지 않는다 (10장 방어책 3)
        대표 = next((t for t in 유형순서 if t in ts), sorted(ts)[0])
        대표정원 = 정원[code].get(대표) or 0
        s = slugify(f["name"], code)
        n = 2
        while s in used:
            s = f"{slugify(f['name'], code)}-{n}"
            n += 1
        used.add(s)
        p = 인력.get(code, {})
        인력합 = sum(p.get(k, 0) for k in
                     ["사회복지사", "간호사", "간호조무사", "물리치료사",
                      "작업치료사", "요양보호사", "영양사"])
        rows.append((code, s, f["name"], f["sido"], f["sigungu"], f["dong"],
                     f["주소"], f["지정일자"],
                     p.get("요양보호사", 0), p.get("간호사", 0), p.get("간호조무사", 0),
                     p.get("사회복지사", 0), p.get("물리치료사", 0), p.get("작업치료사", 0),
                     p.get("의사_촉탁", 0), p.get("영양사", 0),
                     인력합, 대표, 대표정원))
        for t in ts:
            trows.append((code, t, 정원[code].get(t) or 0))

    con.executemany(f"INSERT INTO facility VALUES ({','.join('?'*19)})", rows)
    con.executemany("INSERT INTO facility_type VALUES (?,?,?)", trows)
    con.commit()

    print(f"{시군구} 기관 {len(rows)}곳 저장")
    for t, c in con.execute(
            "SELECT type, COUNT(*) FROM facility_type GROUP BY type ORDER BY COUNT(*) DESC"):
        print(f"  {t:<14}{c:>5}")
    con.close()


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "군산")
